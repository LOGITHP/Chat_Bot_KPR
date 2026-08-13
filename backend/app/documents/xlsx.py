import openpyxl
from typing import List
from app.documents.base import BaseParser, DocumentChunk

class XLSXParser(BaseParser):
    def parse(self, file_path: str) -> List[DocumentChunk]:
        chunks = []
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # Try to extract headers from the first row
                headers = []
                for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                    if row_idx == 0:
                        headers = [str(cell) if cell is not None else f"Col_{i}" for i, cell in enumerate(row)]
                        continue
                    
                    row_data = []
                    for i, cell in enumerate(row):
                        if cell is not None:
                            header = headers[i] if i < len(headers) else f"Col_{i}"
                            row_data.append(f"{header}={cell}")
                    
                    if row_data:
                        text = f"Sheet: {sheet_name}, Row {row_idx + 1}: " + ", ".join(row_data)
                        chunks.append(DocumentChunk(
                            text=text,
                            sheet=sheet_name,
                            row=row_idx + 1
                        ))
            return chunks
        except Exception as e:
            print(f"Error parsing XLSX: {e}")
            return []
